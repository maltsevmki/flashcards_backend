from pathlib import Path
from typing import Union, Optional, List

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from flashcard_importer.models.flashcard import Flashcard, ImportResult, CardType
from flashcard_importer.parsers.base import BaseParser, ParserFactory
from flashcard_importer.utils.validators import CardValidator


class XlsxParser(BaseParser):
    """Parser for Excel (.xlsx) files."""
    
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']
    
    def __init__(self, file_path: Union[str, Path], sheet_name: str = None, has_header: bool = True):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")
        super().__init__(file_path)
        self._sheet_name = sheet_name
        self._has_header = has_header
        self.settings = {}
        self.column_mapping = {}
    
    def detect_settings(self) -> dict:
        """Detect Excel file settings."""
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        
        sheet_names = wb.sheetnames
        active_sheet = self._sheet_name or sheet_names[0]
        
        ws = wb[active_sheet]
        
        # Get first row to check for headers
        first_row = []
        for cell in ws[1]:
            first_row.append(str(cell.value).lower() if cell.value else '')
        
        # Detect headers by checking for common column names
        header_keywords = [
            'front', 'back', 'question', 'answer', 'term', 'definition',
            'deck', 'tags', 'category', 'word', 'meaning'
        ]
        has_header = any(kw in ' '.join(first_row) for kw in header_keywords)
        
        wb.close()
        
        self.settings = {
            'sheet_names': sheet_names,
            'active_sheet': active_sheet,
            'has_header': self._has_header if self._has_header is not None else has_header,
            'first_row': first_row
        }
        
        return self.settings
    
    def set_column_mapping(self, front: int = 0, back: int = 1, deck: int = None, tags: int = None):
        """Set which columns contain which data (0-based index)."""
        self.column_mapping = {
            'front': front,
            'back': back,
            'deck': deck,
            'tags': tags
        }
    
    def get_sheet_names(self) -> List[str]:
        """Get list of all sheet names in the workbook."""
        wb = load_workbook(self.file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    
    def parse(self, sheet_name: str = None) -> ImportResult:
        """Parse the Excel file and return flashcards."""
        result = ImportResult(source_file=str(self.file_path))
        
        if not self.settings:
            self.detect_settings()
        
        active_sheet = sheet_name or self._sheet_name or self.settings.get('active_sheet')
        has_header = self.settings.get('has_header', True)
        
        if not self.column_mapping:
            self.column_mapping = {'front': 0, 'back': 1, 'deck': None, 'tags': None}
        
        result.deck_detected = self.column_mapping.get('deck') is not None
        
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb[active_sheet]
        
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        
        if not rows:
            result.add_error("Empty worksheet")
            return result
        
        # Detect columns from header if present
        start_row = 0
        if has_header and rows:
            self._detect_columns_from_header(rows[0])
            start_row = 1
            result.deck_detected = self.column_mapping.get('deck') is not None
        
        for row_num, row in enumerate(rows[start_row:], start=start_row + 1):
            # Skip empty rows
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            card = self._parse_row(row, row_num, result)
            if card:
                result.add_card(card)
        
        if not result.cards:
            result.add_error("No valid cards found in file")
        
        return result
    
    def _detect_columns_from_header(self, header_row: tuple):
        """Detect column mapping from header names."""
        if not header_row:
            return
        
        header_lower = [str(h).lower().strip() if h else '' for h in header_row]
        
        front_names = ['front', 'question', 'q', 'term', 'word']
        back_names = ['back', 'answer', 'a', 'definition', 'meaning']
        deck_names = ['deck', 'category', 'collection', 'group', 'folder']
        tag_names = ['tags', 'tag', 'labels', 'keywords']
        
        for i, name in enumerate(header_lower):
            if name in front_names:
                self.column_mapping['front'] = i
            elif name in back_names:
                self.column_mapping['back'] = i
            elif name in deck_names:
                self.column_mapping['deck'] = i
            elif name in tag_names:
                self.column_mapping['tags'] = i
    
    def _parse_row(self, row: tuple, row_num: int, result: ImportResult) -> Optional[Flashcard]:
        """Parse a single row into a Flashcard."""
        
        front_col = self.column_mapping.get('front', 0)
        back_col = self.column_mapping.get('back', 1)
        deck_col = self.column_mapping.get('deck')
        tags_col = self.column_mapping.get('tags')
        
        # Check we have enough columns
        if len(row) <= max(front_col, back_col):
            result.add_error(f"Row {row_num}: Not enough columns")
            return None
        
        # Get cell values, handling None
        front_val = row[front_col]
        back_val = row[back_col]
        
        front = str(front_val).strip() if front_val is not None else ''
        back = str(back_val).strip() if back_val is not None else ''
        
        # Validate content
        is_valid, error = CardValidator.validate_content(front)
        if not is_valid:
            result.add_error(f"Row {row_num}: Front - {error}")
            return None
        
        is_valid, error = CardValidator.validate_content(back)
        if not is_valid:
            result.add_error(f"Row {row_num}: Back - {error}")
            return None
        
        # Extract optional fields
        deck_name = None
        tags = []
        
        if deck_col is not None and deck_col < len(row) and row[deck_col]:
            deck_name = str(row[deck_col]).strip() or None
        
        if tags_col is not None and tags_col < len(row) and row[tags_col]:
            tag_string = str(row[tags_col]).strip()
            if tag_string:
                if ',' in tag_string:
                    tags = [t.strip() for t in tag_string.split(',') if t.strip()]
                else:
                    tags = tag_string.split()
        
        # Detect card type
        card_type = CardType.CLOZE if CardValidator.is_cloze_card(front) else CardType.BASIC
        
        return Flashcard(
            front=front,
            back=back,
            deck_name=deck_name,
            tags=tags,
            card_type=card_type,
            source_file=str(self.file_path),
            html_enabled=CardValidator.contains_html(front) or CardValidator.contains_html(back)
        )


# Register the parser
if OPENPYXL_AVAILABLE:
    ParserFactory.register('.xlsx', XlsxParser)
    ParserFactory.register('.xls', XlsxParser)

